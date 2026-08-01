"""Os vocabulários fechados, medidos: 17 tipologias, 169 minerais, e as armadilhas tratadas.

Mesma divisão do `test_refs.py`, pelos mesmos motivos:

1. **Positivos** — travam o snapshot. Vocabulário que mude tem de quebrar aqui, de propósito.
2. **Negativos** — regravam o XLSX em `tmp_path` com **um** defeito por vez e exigem
   `RefLoadError` com a mensagem que nomeia o defeito. Casar só o tipo deixaria a invariante certa
   passar a ser pega pela errada depois de um refactor.
3. **Controle da desingularização e da flag `data_only`** — os dois números do "e se não tivesse".
   São o que impede que alguém remova qualquer uma das duas achando que é enfeite.

Sobre regravar o XLSX: `openpyxl` escreve fórmula **sem** o valor em cache. Isso é ruim para os
negativos em geral (toda cópia dispararia a guarda de cache de B4.2) e é exatamente o que se quer
para *um* deles — por isso `_planilha()` neutraliza `D20` por padrão e `manter_formula=True`
produz o caso da guarda.
"""

from __future__ import annotations

import collections
import re
from pathlib import Path
from typing import Any, Callable

import openpyxl
import pytest
import yaml

from common.dbf import read_dbf
from common.text import fold
from research_pipeline import MAPPING_PATH, REPO_ROOT
from research_pipeline.errors import RefLoadError
from research_pipeline.vocab import (
    RUIDO_SUBSTANCIA,
    build_substancia_index,
    chave_substancia,
    e_sentinela,
    load_minerais,
    load_tipologias,
    separar_uso,
    split_substancias,
    substancias_ambiguas,
)

MAPEAMENTO = yaml.safe_load(MAPPING_PATH.read_text(encoding="utf-8"))
SPEC_TIPOLOGIAS = MAPEAMENTO["tipologias"]
SPEC_MINERAIS = MAPEAMENTO["minerais"]
ABA = SPEC_TIPOLOGIAS["aba"]
ORIGEM_XLSX = REPO_ROOT / SPEC_TIPOLOGIAS["path"]
DBF_IBGE = REPO_ROOT / "data_source" / "Malha municipal IBGE-BA" / "BA_Municipios_2025.dbf"

CODIGOS_63 = {
    "B1.1.1", "B1.1.2", "B1.1.3", "B1.2.1",
    "B2.1", "B2.2",
    "B3.1", "B3.2", "B3.3", "B3.4", "B3.5",
    "B4.1", "B4.2", "B4.3", "B4.4", "B4.5", "B4.6",
}
"""O vocabulário fechado do §6.3, escrito à mão de propósito: é o contrato do AC4."""

AMBIGUAS_MEDIDAS = {
    "calcita": ("B4.3", "B4.5"),
    "caulim": ("B3.3", "B4.1"),
    "caulinita": ("B4.1", "B4.4"),
    "cianita": ("B2.1", "B4.2"),
    "diatomita": ("B4.1", "B4.5"),
    "feldspato": ("B4.2", "B4.4"),
    "granito": ("B3.4", "B3.5"),
    "moscovita": ("B4.2", "B4.4"),
    "quartzito": ("B3.4", "B4.4"),
    "quartzo": ("B3.5", "B4.2"),
    "selenio": ("B1.1.3", "B1.2.1"),
    "sienito": ("B3.4", "B3.5"),
    "turmalina": ("B2.1", "B4.2"),
}
"""Snapshot congelado. O §6.3 lista **dez**; são treze — ver `test_ambiguas_sao_treze`."""


@pytest.fixture(scope="module")
def carga() -> tuple[dict[str, Any], tuple[str, ...]]:
    return load_tipologias(SPEC_TIPOLOGIAS, REPO_ROOT)


@pytest.fixture(scope="module")
def tipologias(carga: tuple[dict[str, Any], tuple[str, ...]]) -> dict[str, Any]:
    return carga[0]


@pytest.fixture(scope="module")
def avisos(carga: tuple[dict[str, Any], tuple[str, ...]]) -> tuple[str, ...]:
    return carga[1]


@pytest.fixture(scope="module")
def indice(tipologias: dict[str, Any]) -> dict[str, tuple[str, ...]]:
    return build_substancia_index(tipologias)


@pytest.fixture(scope="module")
def minerais() -> tuple[tuple[str, ...], dict[str, str]]:
    return load_minerais(SPEC_MINERAIS, REPO_ROOT)


@pytest.fixture(scope="module")
def chaves_exatas(tipologias: dict[str, Any]) -> set[str]:
    """O índice que existiria sob dobra exata — o braço de controle da desingularização.

    A divisão é reproduzida aqui de propósito, com `fold` no lugar de `chave_substancia`: é a
    única forma de medir o "e se não tivesse" sem manter duas implementações no módulo.
    """
    exatas: set[str] = set()
    for t in tipologias.values():
        cabeca, _ = separar_uso(t.nome)
        for parte in cabeca.split(","):
            for pedaco in re.split(r"\s+e\s+", parte):
                dobrado = fold(pedaco)
                if dobrado and dobrado not in RUIDO_SUBSTANCIA:
                    exatas.add(dobrado)
    return exatas


# --------------------------------------------------------------------------- positivos


def test_dezessete_folhas_e_seis_grupos_fora(tipologias: dict[str, Any]) -> None:
    """`fullmatch` mais TIPOLOGIA não vazia dá exatamente o vocabulário fechado do §6.3."""
    assert set(tipologias) == CODIGOS_63
    # As seis linhas de grupo têm coluna A parecida com código e nenhuma coluna B.
    for grupo in ("B1.1", "B1.2", "B2", "B3", "B4", "B1"):
        assert grupo not in tipologias


def test_b42_tem_porte_nulo_e_dois_avisos(tipologias: dict[str, Any], avisos: tuple[str, ...]) -> None:
    """A armadilha do §6.3: `None` nas duas faixas, **nunca `0`**, e um aviso por coluna."""
    b42 = tipologias["B4.2"]
    assert b42.porte_pequeno is None
    assert b42.porte_medio is None
    assert b42.porte_grande == ">= 200.000"
    assert avisos == (
        "tipologia_porte_ausente:B4.2:porte_pequeno",
        "tipologia_porte_ausente:B4.2:porte_medio",
    )


def test_nenhuma_outra_folha_tem_porte_nulo(tipologias: dict[str, Any]) -> None:
    nulos = {
        codigo
        for codigo, t in tipologias.items()
        if None in (t.porte_pequeno, t.porte_medio, t.porte_grande)
    }
    assert nulos == {"B4.2"}


def test_potencial_poluidor(tipologias: dict[str, Any]) -> None:
    histograma = collections.Counter(t.potencial_poluidor for t in tipologias.values())
    assert dict(histograma) == {"A": 10, "M": 7}


def test_unidade_de_porte_e_a_mesma_nas_dezessete(tipologias: dict[str, Any]) -> None:
    """Se um dia deixar de ser, comparar porte entre tipologias vira comparação de unidades."""
    assert {t.unidade_porte for t in tipologias.values()} == {"Produção Bruta de Minério (t/Ano)"}


def test_classes_batem_com_a_matriz_do_art_109(avisos: tuple[str, ...]) -> None:
    """A conferência contra a aba 2: 0 divergências, logo nenhum aviso de classe."""
    assert [a for a in avisos if a.startswith("tipologia_classe_divergente")] == []


def test_indice_de_substancias(indice: dict[str, tuple[str, ...]]) -> None:
    assert len(indice) == 128
    assert indice["granito"] == ("B3.4", "B3.5")
    assert indice["amianto"] == ("B4.6",)
    assert all(codigos == tuple(sorted(codigos)) for codigos in indice.values())


def test_ambiguas_sao_treze(indice: dict[str, tuple[str, ...]]) -> None:
    """São 13, e o §6.3 diz 10. As duas que faltavam lá têm causas independentes.

    `caulim` estava escondido atrás da chave-lixo `"caulim dentre outros"`, produzida por um corte
    de cauda que exigia vírgula antes de `Dentre` — B4.1 escreve sem vírgula. E a colisão importa:
    B3.3 é Caulim sozinho, poluidor **A**, Classe 4/5/6; B4.1 é Caulim junto das argilas, poluidor
    **M**, Classe 2/3/5.

    `quartzo` e `quartzito` só não colidiam por causa do `s` do plural — ver
    `test_desingularizar_funde_exatamente_dois_pares`.
    """
    assert substancias_ambiguas(indice) == AMBIGUAS_MEDIDAS
    assert "caulim dentre outros" not in indice


def test_uso_so_onde_e_informativo(tipologias: dict[str, Any]) -> None:
    """As 5 caudas que o prompt de desambiguação do patch 9 renderiza."""
    com_uso = {codigo for codigo, t in tipologias.items() if t.uso}
    assert com_uso == {"B3.4", "B3.5", "B4.2", "B4.3", "B4.4"}
    assert tipologias["B3.5"].uso == "Dentre Outras Utilizadas Para Revestimento"
    # B4.1 traz `"Dentre Outros"` seco e B2.1 `"e outras"`: cauda sem informação de uso.
    assert tipologias["B4.1"].uso is None
    assert tipologias["B2.1"].uso is None


def test_cauda_cortada_antes_de_dividir() -> None:
    """A cauda de B3.4 contém ` e `; dividir primeiro produziria substância inventada."""
    substancias = split_substancias(
        "Basalto, Calcários, Gnaisses, Granitos, Granulitos, Metarenitos, Quartzitos, Sienitos, "
        "Dentre Outras Utilizadas Para a Produção de Agregados e Beneficiamento Associado (Britamento)"
    )
    assert substancias == (
        "basalto", "calcario", "gnaisse", "granito",
        "granulito", "metarenito", "quartzito", "sienito",
    )
    assert not any("beneficiamento" in s or "agregado" in s for s in substancias)


def test_cauda_sem_virgula_nao_gruda_na_substancia() -> None:
    """B4.1 escreve `"…, Ilita, Caulim Dentre Outros"`. É de onde o `caulim` sumia."""
    assert split_substancias("Argilas, Caulinita, Diatomita, Ilita, Caulim Dentre Outros") == (
        "argila", "caulinita", "diatomita", "ilita", "caulim",
    )


def test_minerais_do_sigmine(minerais: tuple[tuple[str, ...], dict[str, str]]) -> None:
    valores, indice_minerais = minerais
    assert len(valores) == 169
    assert len(indice_minerais) == 169
    assert indice_minerais["granito p/ revestimento"] == "GRANITO P/ REVESTIMENTO"
    linhas = read_dbf(REPO_ROOT / SPEC_MINERAIS["path"])
    assert len(linhas) == 31_858
    assert len(linhas[0]) == 12


# ------------------------------------------------------- controles: os dois "e se não tivesse"


def test_data_only_true_e_carga_estrutural() -> None:
    """O achado que o §6.3 descreve errado, travado nos dois sentidos.

    O §6.3 diz que `#ERROR!` está gravado como *shared string* (`t="s"`). Está como `t="str"`:
    célula de **fórmula** com resultado em cache. A consequência prática é maior que a etiqueta —
    com o default do openpyxl a célula devolve o texto da fórmula, a sentinela não dispara, e um
    porte-lixo entra como válido. Este teste é o que impede alguém de "limpar" a flag.
    """
    formulas = openpyxl.load_workbook(ORIGEM_XLSX, data_only=False)[ABA]
    assert formulas["A20"].value == "B4.2"  # a coordenada não é chute
    assert formulas["D20"].data_type == "f"
    bruto = formulas["D20"].value
    assert "#ERROR!" not in bruto
    assert not e_sentinela(bruto)  # <- o porte-lixo que entraria como válido

    valores = openpyxl.load_workbook(ORIGEM_XLSX, data_only=True)[ABA]
    assert valores["D20"].value == "#ERROR!"
    assert e_sentinela(valores["D20"].value)


def test_desingularizar_funde_exatamente_dois_pares(
    chaves_exatas: set[str], indice: dict[str, tuple[str, ...]]
) -> None:
    """130 chaves sob dobra exata, 128 sob `chave_substancia` — e a diferença são só estes dois."""
    assert len(chaves_exatas) == 130
    assert len(indice) == 128
    fundidas = {
        chave_substancia(c)
        for c in chaves_exatas
        if chave_substancia(c) != c and chave_substancia(c) in chaves_exatas
    }
    assert fundidas == {"quartzo", "quartzito"}
    # Sob dobra exata as quatro resolveriam sozinhas, silenciosamente e metade das vezes errado.
    for singular, plural in (("quartzo", "quartzos"), ("quartzito", "quartzitos")):
        assert {singular, plural} <= chaves_exatas


def test_cruzamento_com_o_sigmine_e_parcial(
    indice: dict[str, tuple[str, ...]],
    chaves_exatas: set[str],
    minerais: tuple[tuple[str, ...], dict[str, str]],
) -> None:
    """69 dos 169, e seriam 56 sem a desingularização. Os dois números no mesmo teste.

    Os vocabulários são largamente disjuntos por natureza — o SIGMINE nomeia minério e rocha, o
    Anexo IV nomeia elemento e mineral. Fica travado para que o patch 9 não seja desenhado
    supondo sobreposição que não existe.
    """
    valores, indice_minerais = minerais
    assert sum(1 for chave in indice_minerais if chave in indice) == 69
    assert sum(1 for valor in valores if fold(valor) in chaves_exatas) == 56


@pytest.mark.parametrize(
    "texto,esperado",
    [
        ("Granitos", "granito"),
        ("Quartzos", "quartzo"),
        ("Ágata", "agata"),
        ("Minerais de Borato", "minerais de borato"),  # só a última palavra perde o `s`
        ("Sílex", "silex"),
        ("Gás", "gas"),  # 3 letras: curta demais para o `s` ser plural
        ("Lápis-Lazúli", "lapis lazuli"),
    ],
)
def test_chave_substancia(texto: str, esperado: str) -> None:
    assert chave_substancia(texto) == esperado


# --------------------------------------------------------------------------- negativos


def _planilha(
    tmp_path: Path, mutacao: Callable[[Any], None] | None = None, manter_formula: bool = False
) -> Path:
    """Regrava o XLSX real em `tmp_path` com um defeito. Ver o docstring do módulo."""
    wb = openpyxl.load_workbook(ORIGEM_XLSX, data_only=False)
    if not manter_formula:
        wb[ABA]["D20"] = "#ERROR!"  # openpyxl não regrava o cache; sem isto toda cópia falharia
    if mutacao is not None:
        mutacao(wb)
    destino = tmp_path / SPEC_TIPOLOGIAS["path"]
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino


def _remover_uma_folha(wb: Any) -> None:
    wb[ABA].delete_rows(24)  # B4.6 Amianto


def _codigo_inventado(wb: Any) -> None:
    """Linha-folha nova e **completa**, para que a falha seja a contagem e não um campo vazio.

    `append` e não escrita por coordenada: as linhas logo abaixo das 17 estão dentro de faixas
    mescladas, e `MergedCell.value` é somente-leitura.
    """
    ws = wb[ABA]
    modelo = [ws.cell(row=24, column=coluna).value for coluna in range(1, 11)]
    modelo[0] = "B9.9"
    ws.append(modelo)


def _renomear_coluna_codigo(wb: Any) -> None:
    wb[ABA]["A1"] = "COD"


def _potencial_invalido(wb: Any) -> None:
    wb[ABA]["G4"] = "X"


def _tipologia_vazia(wb: Any) -> None:
    wb[ABA]["B4"] = None


def _unidade_vazia(wb: Any) -> None:
    wb[ABA]["C4"] = None


def _classe_vazia(wb: Any) -> None:
    wb[ABA]["H4"] = None


def _aba_renomeada(wb: Any) -> None:
    wb[ABA].title = "Outra Aba"


def _matriz_furada(wb: Any) -> None:
    wb[SPEC_TIPOLOGIAS["aba_matriz"]]["C3"] = None


CORRUPCOES = [
    (_remover_uma_folha, "16 tipologias-folha, esperado 17"),
    (_codigo_inventado, "18 tipologias-folha, esperado 17"),
    (_renomear_coluna_codigo, "coluna 'CÓDIGO' ausente"),
    (_potencial_invalido, "POTENCIAL POLUIDOR 'X' fora de"),
    (_tipologia_vazia, "B1.1.1 sem TIPOLOGIA"),
    (_unidade_vazia, "UNIDADE DE MEDIDA DE PORTE vazia"),
    (_classe_vazia, "CLASSE (pequeno) vazia"),
    (_aba_renomeada, "aba 'Divisão B - Mineração' ausente"),
    (_matriz_furada, "matriz porte x potencial poluidor tem 8 células"),
]


@pytest.mark.parametrize(
    "mutacao,esperado", CORRUPCOES, ids=[m.__name__.lstrip("_") for m, _ in CORRUPCOES]
)
def test_planilha_corrompida_falha_alto(
    tmp_path: Path, mutacao: Callable[[Any], None], esperado: str
) -> None:
    _planilha(tmp_path, mutacao)
    with pytest.raises(RefLoadError) as capturado:
        load_tipologias(SPEC_TIPOLOGIAS, tmp_path)
    assert esperado in str(capturado.value)


def test_formula_sem_valor_em_cache(tmp_path: Path) -> None:
    """A guarda que troca `None` silencioso por falha alta.

    Um XLSX regravado por ferramenta que não avalia fórmula devolve `None` na célula — que a
    camada de porte trataria como "faixa ausente", emitiria o aviso de sempre e seguiria em
    frente, escondendo que o arquivo mudou.
    """
    _planilha(tmp_path, manter_formula=True)
    with pytest.raises(RefLoadError, match="sem valor em cache"):
        load_tipologias(SPEC_TIPOLOGIAS, tmp_path)


def test_sentinela_nova_vira_aviso_e_nao_erro(tmp_path: Path) -> None:
    """Sentinela é defeito **da publicação**, que não vai ser corrigida: aviso, nunca erro."""
    _planilha(tmp_path, lambda wb: wb[ABA].__setitem__("D4", "#N/A"))
    tipologias, avisos = load_tipologias(SPEC_TIPOLOGIAS, tmp_path)
    assert tipologias["B1.1.1"].porte_pequeno is None
    assert "tipologia_porte_ausente:B1.1.1:porte_pequeno" in avisos


def test_classe_divergente_vira_aviso_e_nao_erro(tmp_path: Path) -> None:
    """A matriz é a regra geral do Art. 109; a folha é a publicação. Discordar não impede carga."""
    _planilha(tmp_path, lambda wb: wb[ABA].__setitem__("H4", "Classe 1"))
    tipologias, avisos = load_tipologias(SPEC_TIPOLOGIAS, tmp_path)
    assert tipologias["B1.1.1"].classe_pequeno == "Classe 1"
    assert "tipologia_classe_divergente:B1.1.1:classe_pequeno" in avisos


def test_planilha_ausente(tmp_path: Path) -> None:
    with pytest.raises(RefLoadError, match="planilha de tipologias ausente"):
        load_tipologias(SPEC_TIPOLOGIAS, tmp_path)


def test_xlsx_ilegivel(tmp_path: Path) -> None:
    destino = tmp_path / SPEC_TIPOLOGIAS["path"]
    destino.parent.mkdir(parents=True)
    destino.write_bytes(b"nao sou um zip")
    with pytest.raises(RefLoadError, match="XLSX ilegível"):
        load_tipologias(SPEC_TIPOLOGIAS, tmp_path)


def test_dbf_ausente(tmp_path: Path) -> None:
    with pytest.raises(RefLoadError, match="DBF de minerais ausente"):
        load_minerais(SPEC_MINERAIS, tmp_path)


def test_dbf_sem_a_coluna_mapeada() -> None:
    """A malha do IBGE é um DBF real e válido — só não tem `SUBS`."""
    spec = {**SPEC_MINERAIS, "path": str(DBF_IBGE.relative_to(REPO_ROOT))}
    with pytest.raises(RefLoadError, match="coluna 'SUBS' ausente"):
        load_minerais(spec, REPO_ROOT)


def test_contagem_de_minerais_divergente() -> None:
    with pytest.raises(RefLoadError, match="169 valores distintos em 'SUBS', esperado 168"):
        load_minerais({**SPEC_MINERAIS, "esperados": 168}, REPO_ROOT)


def test_contagem_de_tipologias_divergente() -> None:
    with pytest.raises(RefLoadError, match="17 tipologias-folha, esperado 16"):
        load_tipologias({**SPEC_TIPOLOGIAS, "esperadas": 16}, REPO_ROOT)
