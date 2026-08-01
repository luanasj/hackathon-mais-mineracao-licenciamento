"""Vocabulários fechados: as 17 tipologias do Anexo IV e os 169 minerais do SIGMINE.

Extensão do critério de aceite 8 (`refs.py`) ao vocabulário: o AC4 exige que todo
`tipologia_codigo` esteja no conjunto fechado de 17, e não há como medir isso sem carregar o
conjunto **antes** de gastar. Uma planilha corrompida descoberta no nó `normalize` já custou o
relatório de US$ 3.

Quatro coisas que o formato real dos arquivos impôs e que não são óbvias no `GOAL.md` §6.3:

1. **`data_only=True` é carga estrutural.** A célula PORTE PEQUENO de B4.2 é uma **fórmula com
   resultado em cache**, `<c t="str"><f> 20.000 &lt; 200.000 …</f><v>#ERROR!</v></c>` — não uma
   *shared string* `t="s"`, como o §6.3 afirma, nem uma célula de erro `t="e"`. Com o default do
   openpyxl a célula devolve o **texto da fórmula**, que não contém `#ERROR!` e não casa
   `faixa não expressa`: a sentinela não dispara e o porte-lixo entra como válido. Detectar por
   sentinela de texto é necessário e **não é suficiente**.
2. **A fórmula mostra o que a corrupção comeu.** Comparando com B4.6
   (`< 20.000` / `>= 20.000 < 300.000` / `>= 300.000`), a faixa ` 20.000 < 200.000` que sobrou em
   PEQUENO é a faixa **MÉDIO** deslocada, e o PEQUENO é o que a publicação oficial omitiu. Ou
   seja: não é um erro de cálculo, é dado deslocado, e não se sabe para onde. Confirma a decisão
   travada do §6.3 — `None` nas duas colunas e um aviso por coluna, **nunca `0`** — por um motivo
   melhor do que "a célula deu erro".
3. **As substâncias ambíguas são 13, não as 10 do §6.3.** Duas causas somadas: `caulim` estava
   escondido atrás de uma chave-lixo (ver `CAUDA_RE`), e `quartzo`/`quartzito` só não colidiam
   por causa do `s` do plural (ver `chave_substancia`). O conjunto é **derivado**, nunca escrito
   à mão — o teste congela o snapshot, o código não.
4. **Os dois vocabulários são largamente disjuntos.** Só 69 dos 169 `SUBS` existem no índice do
   Anexo IV: o SIGMINE nomeia minério e rocha (`MINERIO DE FERRO`, `MIGMATITO`), o Anexo IV nomeia
   elemento e mineral. Não é defeito de nenhum dos dois — são duas buscas distintas no patch 9,
   `substancia_raw` contra o índice e `mineral` contra os 169 — mas fica medido para que ninguém
   desenhe o patch 9 supondo sobreposição que não existe.

Divisão de responsabilidade com `refs.py`: aqui célula malformada **conhecida** (as sentinelas)
vira aviso, porque a publicação oficial é a fonte e não vai ser corrigida; qualquer outra coisa
fora de forma é `RefLoadError`, porque significa que o arquivo mudou.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, get_args

import openpyxl

from common.dbf import read_dbf
from common.text import fold
from research_pipeline.errors import RefLoadError
from research_pipeline.schemas import PotencialPoluidor

LEAF_CODE_RE = re.compile(r"B\d+(?:\.\d+){1,2}")
"""Usado com `.fullmatch()`, e isso é carga estrutural.

Seis linhas da planilha são de grupo e trazem na coluna A `"B1.1 Minerais metálicos"`,
`"B1.2 Minerais Não Metálicos"` e afins. `match()` ou `startswith("B")` engoliria as seis como
tipologia. Fullmatch mais coluna TIPOLOGIA não vazia dá exatamente 17.
"""

SENTINELAS_ERRO = frozenset({"#ERROR!", "#REF!", "#N/A", "#VALUE!", "#DIV/0!"})
SENTINELA_TEXTO_RE = re.compile(r"faixa n[ãa]o expressa", re.IGNORECASE)
"""As duas formas em que a publicação oficial registra "esta faixa não existe"."""

CAUDA_RE = re.compile(
    r"\s*,?\s*(?:Dentre\s+Out\w*|Para\s+Manufatura|Para\s+Produ\w*)|\s+e\s+outras\b",
    re.IGNORECASE,
)
"""Corta a cauda de uso do nome da tipologia. Duas decisões dentro dela:

**Corta antes de dividir**, porque a cauda de B3.4 contém ` e `
(`"…Agregados e Beneficiamento Associado (Britamento)"`) e dividir primeiro produziria
`"beneficiamento associado (britamento)"` como se fosse substância.

**A vírgula é opcional** (`,?`). B4.1 escreve `"…, Ilita, Caulim Dentre Outros"` **sem** vírgula
antes de `Dentre`. Exigir a vírgula deixa a chave-lixo `"caulim dentre outros"` no índice e
esconde a colisão real: `caulim` está em B3.3 sozinho (poluidor **A**, Classe 4/5/6) e em B4.1
junto das argilas (poluidor **M**, Classe 2/3/5). É de onde vem o `10` do §6.3, que é 11.
"""

CAUDAS_VAZIAS = frozenset({"dentre outros", "dentre outras", "e outras", "e outros"})
"""Caudas sem informação de uso: viram `uso=None` em vez de texto inútil no prompt do patch 9."""

RUIDO_SUBSTANCIA = frozenset({"etc", "etc.", "outras", "outros"})

POTENCIAIS_VALIDOS = get_args(PotencialPoluidor)
"""O §6.3 admite os três. Medido: as 17 folhas usam só `M` (7) e `A` (10) — a matriz do Art. 109
tem coluna `P` e o `Literal` acompanha a matriz, não o snapshot. Derivado do literal de
`schemas.py` em vez de reescrito: a checagem em tempo de carga e a do schema têm de ser o mesmo
conjunto, sempre."""

PORTES = ("pequeno", "medio", "grande")


@dataclass(frozen=True, slots=True)
class Tipologia:
    """Uma das 17 folhas da Divisão B.

    Portes ficam **string crua** (`"< 300.000"`, `">= 300.000 < 1.500.000"`): o produto do §8 não
    os usa numericamente, e um parser de faixa seria superfície sem cliente que ainda teria de
    decidir o que fazer com a faixa deslocada de B4.2.

    `uso` é a cauda cortada do nome, quando informativa — é o texto que distingue B3.4 (britagem)
    de B3.5 (revestimento) e que o prompt de desambiguação do patch 9 renderiza por substância
    ambígua. Fica `None` nas 12 folhas sem cauda útil.
    """

    codigo: str
    nome: str
    unidade_porte: str
    porte_pequeno: str | None
    porte_medio: str | None
    porte_grande: str | None
    potencial_poluidor: PotencialPoluidor
    classe_pequeno: str
    classe_medio: str
    classe_grande: str
    substancias: tuple[str, ...]
    uso: str | None


def chave_substancia(texto: str) -> str:
    """Chave de comparação de substância: `fold()` mais o plural da última palavra.

    Os dois lados usam esta função — o índice e a consulta do patch 9 — exatamente como `fold` faz
    para município e consórcio no patch 6.

    O `s` final sai só da **última** palavra, e só com mais de 3 letras: `"Minerais de Borato"`
    tem de continuar `minerais de borato` e não virar `minerai de borato`.

    Por que desingularizar: o Anexo IV escreve a mesma substância no plural numa folha e no
    singular noutra. `Quartzo` está em B4.2 (vidro/óptica) e `Quartzos` em B3.5 (revestimento);
    `Quartzito` em B4.4 (industrial) e `Quartzitos` em B3.4 (britagem). Sob dobra exata um
    relatório que escreva `"Quartzo"` resolveria silenciosamente para B4.2 e nunca chegaria ao
    LLM — o mesmo defeito do Granito, sem o alarme. Medido nas 130 chaves cruas, isto funde
    exatamente esses 2 pares e nenhum outro; nos 169 `SUBS` não funde nada.

    >>> chave_substancia("Granitos")
    'granito'
    >>> chave_substancia("Minerais de Borato")
    'minerais de borato'
    """
    palavras = fold(texto).split()
    if palavras and len(palavras[-1]) > 3 and palavras[-1].endswith("s"):
        palavras[-1] = palavras[-1][:-1]
    return " ".join(palavras)


def separar_uso(nome: str) -> tuple[str, str | None]:
    """Divide o nome da tipologia em (lista de substâncias, cauda de uso)."""
    achado = CAUDA_RE.search(nome)
    if achado is None:
        return nome.strip(), None
    cabeca = nome[: achado.start()].strip().rstrip(",").strip()
    cauda = nome[achado.start() :].strip().lstrip(",").strip()
    return cabeca, None if fold(cauda) in CAUDAS_VAZIAS else cauda


def split_substancias(nome: str) -> tuple[str, ...]:
    """Nome da tipologia -> chaves de substância, sem a cauda de uso e sem repetição."""
    cabeca, _ = separar_uso(nome)
    chaves: list[str] = []
    for parte in cabeca.split(","):
        for pedaco in re.split(r"\s+e\s+", parte):
            if not fold(pedaco) or fold(pedaco) in RUIDO_SUBSTANCIA:
                continue
            chaves.append(chave_substancia(pedaco))
    return tuple(dict.fromkeys(chaves))


def e_sentinela(valor: Any) -> bool:
    """`True` para as formas em que a publicação registra faixa inexistente."""
    if not isinstance(valor, str):
        return False
    return valor.strip() in SENTINELAS_ERRO or bool(SENTINELA_TEXTO_RE.search(valor))


def build_substancia_index(tipologias: dict[str, Tipologia]) -> dict[str, tuple[str, ...]]:
    """Índice substância -> tipologias. Muitos-para-muitos **por construção**."""
    indice: dict[str, list[str]] = {}
    for codigo in sorted(tipologias):
        for substancia in tipologias[codigo].substancias:
            indice.setdefault(substancia, []).append(codigo)
    return {chave: tuple(codigos) for chave, codigos in sorted(indice.items())}


def substancias_ambiguas(indice: dict[str, tuple[str, ...]]) -> dict[str, tuple[str, ...]]:
    """As substâncias que mapeiam para mais de uma tipologia. **Derivado, nunca escrito à mão.**

    São as únicas linhas que o patch 9 manda ao LLM para desambiguar por uso, e a seção do prompt
    é renderizada a partir daqui — uma frase fixa sobre Granito deixaria as outras doze sem
    instrução.
    """
    return {chave: codigos for chave, codigos in indice.items() if len(codigos) > 1}


def _indice_colunas(ws: Any, colunas: dict[str, str], linha: int, rotulo: str) -> dict[str, int]:
    """Localiza as colunas pelo **texto** do cabeçalho, nunca por posição."""
    cabecalhos: dict[str, int] = {}
    for celula in ws[linha]:
        if isinstance(celula.value, str) and celula.value.strip():
            cabecalhos[celula.value.strip()] = celula.column
    indice: dict[str, int] = {}
    for destino, titulo in colunas.items():
        if titulo not in cabecalhos:
            raise RefLoadError(
                f"{rotulo}: coluna {titulo!r} ausente (-> {destino}); "
                f"o cabeçalho traz {sorted(cabecalhos)}"
            )
        indice[destino] = cabecalhos[titulo]
    return indice


def _valor_da_celula(ws_valores: Any, ws_formulas: Any, linha: int, coluna: int, rotulo: str) -> Any:
    """Lê a célula pelo valor, e falha alto se for fórmula sem cache ou fórmula desconhecida.

    O Anexo IV é publicação estática: nenhuma célula deveria calcular. A única que calcula é a
    corrupção conhecida de B4.2, e o resultado dela em cache é uma sentinela. Sem esta guarda, um
    XLSX regravado por ferramenta que não avalia fórmula devolveria `None` — que a camada de porte
    trataria como "faixa ausente" e seguiria em frente com um aviso, escondendo o defeito.
    """
    celula_formula = ws_formulas.cell(row=linha, column=coluna)
    valor = ws_valores.cell(row=linha, column=coluna).value
    if celula_formula.data_type != "f":
        return valor
    if valor is None:
        raise RefLoadError(
            f"{rotulo}: célula {celula_formula.coordinate} é fórmula "
            f"({celula_formula.value!r}) sem valor em cache — regrave a planilha com uma "
            "ferramenta que avalie fórmulas, ou substitua a célula pelo texto literal"
        )
    if not e_sentinela(valor):
        raise RefLoadError(
            f"{rotulo}: célula {celula_formula.coordinate} é fórmula "
            f"({celula_formula.value!r}) e o resultado {valor!r} não é sentinela conhecida — "
            "o Anexo IV é publicação estática e não deveria ter célula que calcula"
        )
    return valor


def _porte(valor: Any, codigo: str, campo: str, avisos: list[str]) -> str | None:
    """Faixa de porte, ou `None` com aviso. **Nunca `0`** — ausente não é zero (§6.3)."""
    if valor is None or (isinstance(valor, str) and not valor.strip()) or e_sentinela(valor):
        avisos.append(f"tipologia_porte_ausente:{codigo}:{campo}")
        return None
    return str(valor).strip()


def _load_matriz(wb: Any, aba: str, rotulo: str) -> dict[tuple[str, str], str]:
    """Lê a matriz Porte × Potencial Poluidor -> Classe do Art. 109 (aba 2), não hardcoded."""
    if aba not in wb.sheetnames:
        raise RefLoadError(f"{rotulo}: aba {aba!r} ausente; o arquivo traz {wb.sheetnames}")
    ws = wb[aba]
    potencial_por_coluna: dict[int, str] = {}
    for celula in ws[1]:
        if celula.column == 1 or not isinstance(celula.value, str):
            continue
        achado = re.search(r"\(([PMA])\)", celula.value)
        if achado:
            potencial_por_coluna[celula.column] = achado.group(1)

    matriz: dict[tuple[str, str], str] = {}
    for linha in ws.iter_rows(min_row=2):
        rotulo_porte = linha[0].value
        if not isinstance(rotulo_porte, str) or fold(rotulo_porte) not in PORTES:
            continue
        for celula in linha[1:]:
            potencial = potencial_por_coluna.get(celula.column)
            if potencial and isinstance(celula.value, str):
                matriz[(fold(rotulo_porte), potencial)] = celula.value.strip()

    if len(matriz) != len(PORTES) * len(POTENCIAIS_VALIDOS):
        raise RefLoadError(
            f"{rotulo} [{aba}]: matriz porte x potencial poluidor tem {len(matriz)} células, "
            f"esperado {len(PORTES) * len(POTENCIAIS_VALIDOS)}"
        )
    return matriz


def load_tipologias(spec: dict[str, Any], root: Path) -> tuple[dict[str, Tipologia], tuple[str, ...]]:
    """Carrega as 17 tipologias-folha e devolve `(tipologias, avisos)`.

    `spec` é o bloco `tipologias:` de `ref_mapping.yaml`; `root` é a raiz de resolução do `path:`,
    parametrizada para os testes negativos, igual a `refs.load_reference_data`.
    """
    caminho = root / spec["path"]
    if not caminho.exists():
        raise RefLoadError(f"{caminho}: planilha de tipologias ausente")
    try:
        wb_valores = openpyxl.load_workbook(caminho, data_only=True)
        wb_formulas = openpyxl.load_workbook(caminho, data_only=False)
    except RefLoadError:
        raise
    except Exception as erro:  # zipfile.BadZipFile, KeyError de parte ausente, etc.
        raise RefLoadError(f"{caminho}: XLSX ilegível ({type(erro).__name__}: {erro})") from erro

    aba = spec["aba"]
    if aba not in wb_valores.sheetnames:
        raise RefLoadError(f"{caminho}: aba {aba!r} ausente; o arquivo traz {wb_valores.sheetnames}")
    ws_valores = wb_valores[aba]
    ws_formulas = wb_formulas[aba]

    linha_cabecalho = spec.get("linha_cabecalho", 1)
    colunas = _indice_colunas(ws_valores, spec["colunas"], linha_cabecalho, f"{caminho} [{aba}]")
    matriz = _load_matriz(wb_valores, spec["aba_matriz"], str(caminho))

    tipologias: dict[str, Tipologia] = {}
    avisos: list[str] = []
    for numero in range(linha_cabecalho + 1, ws_valores.max_row + 1):
        bruto = ws_valores.cell(row=numero, column=colunas["codigo"]).value
        nome = ws_valores.cell(row=numero, column=colunas["nome"]).value
        if not isinstance(bruto, str) or not LEAF_CODE_RE.fullmatch(bruto.strip()):
            continue  # linha de grupo, linha vazia ou rodapé — ver LEAF_CODE_RE
        codigo = bruto.strip()
        if not isinstance(nome, str) or not nome.strip():
            raise RefLoadError(f"{caminho} [{aba}] linha {numero}: {codigo} sem TIPOLOGIA")
        if codigo in tipologias:
            raise RefLoadError(f"{caminho} [{aba}] linha {numero}: código {codigo} duplicado")

        rotulo = f"{caminho.name} [{aba}]: {codigo}"
        campos = {
            destino: _valor_da_celula(ws_valores, ws_formulas, numero, coluna, rotulo)
            for destino, coluna in colunas.items()
        }

        potencial = campos["potencial_poluidor"]
        if potencial not in POTENCIAIS_VALIDOS:
            raise RefLoadError(
                f"{rotulo}: POTENCIAL POLUIDOR {potencial!r} fora de {list(POTENCIAIS_VALIDOS)}"
            )
        unidade = campos["unidade_porte"]
        if not isinstance(unidade, str) or not unidade.strip():
            raise RefLoadError(f"{rotulo}: UNIDADE DE MEDIDA DE PORTE vazia")

        classes = {}
        for porte in PORTES:
            valor = campos[f"classe_{porte}"]
            if not isinstance(valor, str) or not valor.strip():
                raise RefLoadError(f"{rotulo}: CLASSE ({porte}) vazia")
            classes[porte] = valor.strip()
            esperada = matriz[(porte, potencial)]
            if classes[porte] != esperada:
                # Aviso e não erro: a matriz é a regra geral do Art. 109, a folha é a publicação.
                avisos.append(f"tipologia_classe_divergente:{codigo}:classe_{porte}")

        nome = nome.strip()
        _, uso = separar_uso(nome)
        tipologias[codigo] = Tipologia(
            codigo=codigo,
            nome=nome,
            unidade_porte=unidade.strip(),
            porte_pequeno=_porte(campos["porte_pequeno"], codigo, "porte_pequeno", avisos),
            porte_medio=_porte(campos["porte_medio"], codigo, "porte_medio", avisos),
            porte_grande=_porte(campos["porte_grande"], codigo, "porte_grande", avisos),
            potencial_poluidor=potencial,
            classe_pequeno=classes["pequeno"],
            classe_medio=classes["medio"],
            classe_grande=classes["grande"],
            substancias=split_substancias(nome),
            uso=uso,
        )

    esperadas = spec["esperadas"]
    if len(tipologias) != esperadas:
        raise RefLoadError(
            f"{caminho} [{aba}]: {len(tipologias)} tipologias-folha, esperado {esperadas} "
            f"({sorted(tipologias)})"
        )
    return tipologias, tuple(avisos)


def load_minerais(spec: dict[str, Any], root: Path) -> tuple[tuple[str, ...], dict[str, str]]:
    """Carrega os 169 `SUBS` do SIGMINE e devolve `(valores, índice chave -> valor original)`."""
    caminho = root / spec["path"]
    if not caminho.exists():
        raise RefLoadError(f"{caminho}: DBF de minerais ausente")
    coluna = spec["coluna"]
    linhas = read_dbf(caminho, encoding=spec.get("encoding", "utf-8"))
    if not linhas:
        raise RefLoadError(f"{caminho}: DBF sem registros")
    if coluna not in linhas[0]:
        raise RefLoadError(f"{caminho}: coluna {coluna!r} ausente; o DBF traz {sorted(linhas[0])}")

    valores = {linha[coluna].strip() for linha in linhas}
    if "" in valores:
        raise RefLoadError(f"{caminho}: coluna {coluna!r} tem registro com valor vazio")
    minerais = tuple(sorted(valores))
    esperados = spec["esperados"]
    if len(minerais) != esperados:
        raise RefLoadError(
            f"{caminho}: {len(minerais)} valores distintos em {coluna!r}, esperado {esperados}"
        )

    indice: dict[str, str] = {}
    for valor in minerais:
        chave = chave_substancia(valor)
        if chave in indice:
            raise RefLoadError(
                f"{caminho}: {valor!r} e {indice[chave]!r} colidem na chave {chave!r} — "
                "o vocabulário de mineral deixou de ser unívoco sob `chave_substancia`"
            )
        indice[chave] = valor
    return minerais, indice


if __name__ == "__main__":
    import yaml

    from research_pipeline import MAPPING_PATH, REPO_ROOT

    mapeamento = yaml.safe_load(MAPPING_PATH.read_text(encoding="utf-8"))
    tipologias, avisos = load_tipologias(mapeamento["tipologias"], REPO_ROOT)
    minerais, indice_minerais = load_minerais(mapeamento["minerais"], REPO_ROOT)

    indice = build_substancia_index(tipologias)
    ambiguas = substancias_ambiguas(indice)
    potenciais = {p: sum(1 for t in tipologias.values() if t.potencial_poluidor == p) for p in POTENCIAIS_VALIDOS}
    casam = sum(1 for chave in indice_minerais if chave in indice)

    print(f"{len(tipologias)} tipologias ({potenciais['A']} A / {potenciais['M']} M / {potenciais['P']} P)")
    print(f"{len(indice)} chaves de substância, {len(ambiguas)} ambíguas")
    for chave, codigos in ambiguas.items():
        print(f"  {chave:12} {' / '.join(codigos)}")
    print(f"{len(minerais)} minerais ({casam} casam com o Anexo IV)")
    print(f"{sum(1 for t in tipologias.values() if t.uso)} tipologias com uso declarado")
    for aviso in avisos:
        print(f"aviso: {aviso}")
    print(f"{len(avisos)} avisos")
